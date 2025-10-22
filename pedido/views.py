from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, View
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Produto, ItemPedido, Pedido, Cliente, PedidoEntregue, PedidoCancelado, Combo
from django.http import HttpResponse
from django.db import transaction 
from django.utils.formats import number_format
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
import pandas as pd 
import json
from django.db.models import Case, When, Value, IntegerField, Min
from django.utils import timezone
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict
from django.core.serializers.json import DjangoJSONEncoder
from collections import Counter
import re
import pytz



class PagInicial(TemplateView):
    template_name = "pag_inicial.html"

@csrf_exempt
def salvar_nome(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        nome = data.get('nome')

        if nome:
            cliente = Cliente.objects.create(nome=nome)
            return JsonResponse({'redirect_url': f'/pedido/{cliente.id}/'})
        
        return JsonResponse({'erro': 'Nome não pode estar vazio'}, status=400)

    return JsonResponse({'erro': 'Método não permitido'}, status=405)

def devolver_estoque_do_item(item):
    produto = item.produto
    quantidade = item.quantidade

    if hasattr(produto, 'combo'):
        combo = produto.combo

        if combo.tipo == 'fixo':
            # devolve todos os produtos fixos do combo
            for combo_item in combo.itens.all():
                prod = combo_item.produto
                prod.quantidade += quantidade
                prod.save()

        elif combo.tipo == 'opcional':
            # devolve os fixos
            for combo_item in combo.itens.all():
                prod = combo_item.produto
                prod.quantidade += quantidade
                prod.save()

            # devolve as escolhas específicas do item
            escolhas_ids = item.escolhas_combo or []
            for escolha_id in escolhas_ids:
                prod_escolha = Produto.objects.get(id=escolha_id)
                prod_escolha.quantidade += 1  # cada escolha é 1 unid
                prod_escolha.save()

    else:
        # Produto comum
        produto.quantidade += quantidade
        produto.save()

def combos_disponiveis():
    combos = Combo.objects.prefetch_related("itens__produto")
    return [combo for combo in combos if combo.disponivel()]

def itempedido_para_json(item):
    produto = item.produto
    dados = {
        "id": item.id,
        "nome": produto.nome,
        "imagem_url": produto.imagem.url if produto.imagem else "",
        "itens": []
    }

    if hasattr(produto, "combo") and produto.combo.tipo == "opcional":
        # Adiciona os produtos fixos do combo
        for combo_item in produto.combo.itens.all():
            p = combo_item.produto
            dados["itens"].append({
                "id": p.id,
                "nome": p.nome,
                "imagem_url": p.imagem.url if p.imagem else ""
            })

        # Adiciona as escolhas feitas
        for escolha_id in item.escolhas_combo or []:
            p = Produto.objects.get(id=escolha_id)
            dados["itens"].append({
                "id": p.id,
                "nome": p.nome,
                "imagem_url": p.imagem.url if p.imagem else ""
            })

    return dados

@require_POST
def detalhes_item_para_remocao(request, item_id):
    try:
        item = ItemPedido.objects.get(id=item_id)
        dados_item = itempedido_para_json(item)
        return JsonResponse({"status": "ok", "item": dados_item})
    except ItemPedido.DoesNotExist:
        return JsonResponse({"status": "erro", "mensagem": "Item não encontrado"})

class PagCliente(ListView):
    template_name = "pag_cliente.html"
    model = Produto  

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        cliente_id = self.kwargs.get('cliente_id')
        cliente = get_object_or_404(Cliente, id=cliente_id)
        context['cliente'] = cliente

        # pega só produtos com estoque
        produtos = Produto.objects.filter(quantidade__gt=0).order_by('nome')

        # pega combos que só têm itens com estoque suficiente
        combos = combos_disponiveis()

        # Ordem desejada
        ordem_desejada = ["Combos", "Caldos", "Salgados", "Doces", "Bebidas"]

        # Agrupar produtos por categoria válida
        categorias_dict = defaultdict(list)
        for produto in produtos:
            if produto.categoria:
                categoria = produto.categoria.strip().capitalize()
                if categoria in ordem_desejada:
                    categorias_dict[categoria].append(produto)

        # Inserir os combos manualmente na categoria "Combos"
        if combos:
            categorias_dict["Combos"] = [c.produto_ptr for c in combos]

        # Criar dict ordenado manualmente
        categorias_produtos = OrderedDict()
        for cat in ordem_desejada:
            if cat in categorias_dict:
                categorias_produtos[cat] = categorias_dict[cat]

        context['categorias_produtos'] = categorias_produtos.items()

        # Opções de combos opcionais
        # pega combos que têm opções (opcional ou fixo_opcional)
        combos_opcionais = Combo.objects.filter(tipo__in=['opcional', 'fixo_opcional'])
        opcoes_json = {}

        for combo in combos_opcionais:
            opcoes = []
            for opcao in combo.opcoes.all():
                produtos = Produto.objects.filter(categoria=opcao.categoria, quantidade__gt=0)
                opcoes.append({
                    'categoria': opcao.categoria,
                    'categoria_nome': dict(Produto.CATEGORIAS).get(opcao.categoria),
                    'produtos': [{
                        'id': p.id,
                        'nome': p.nome,
                        'imagem_url': p.imagem.url if p.imagem else ""
                    } for p in produtos]
                })

            opcoes_json[combo.produto_ptr.id] = opcoes


        context['opcoes_json'] = json.dumps(opcoes_json, cls=DjangoJSONEncoder)
        return context

    def dispatch(self, request, *args, **kwargs):
        cliente_id = self.kwargs.get('cliente_id')
        cliente = get_object_or_404(Cliente, id=cliente_id)

        pedido_existente = Pedido.objects.filter(cliente=cliente).last()

        if pedido_existente and pedido_existente.liberado_para_caixa:
            print("entrou")
            return redirect('pag_final_cliente', pedido_id=pedido_existente.id)

        return super().dispatch(request, *args, **kwargs)
 
    def post(self, request, cliente_id):
        try:
            data = json.loads(request.body)

            metodo_pagamento = data.get("metodo_pagamento")
            tipo_consumo = data.get("tipo_consumo")
            observacoes = data.get("observacoes", "")
            itens = data.get("itens")

            if not metodo_pagamento or not tipo_consumo or not itens:
                return JsonResponse({"status": "erro", "mensagens": ["Dados incompletos."]})

            mensagens_erro = []

            with transaction.atomic():
                cliente = get_object_or_404(Cliente, id=cliente_id)

                # 1. Verifica estoque (continua igual, apenas conferindo)
                for item in itens:
                    produto_id = item.get("produto_id")
                    quantidade = item.get("quantidade", 1)
                    escolhas = item.get("escolhas", [])

                    produto = Produto.objects.select_for_update().get(id=produto_id)

                    if hasattr(produto, 'combo'):
                        combo = produto.combo

                        combo.refresh_from_db()
                        if not combo.disponivel(quantidade):
                            mensagens_erro.append(
                                f'Combo "{combo.produto_ptr.nome}" não tem estoque suficiente para {quantidade} unidade(s).'
                            )

                        if combo.tipo in ['opcional', 'fixo_opcional']:
                            # Para cada unidade pedida, precisamos garantir que exista alguém para cada escolha
                            # Aqui assumimos que cliente envia 'escolhas' como lista adequada (ver front-end)
                            # Se 'escolhas' vier agrupado (grupos por combo) faça a validação por grupo
                            if isinstance(escolhas, list) and any(isinstance(e, list) for e in escolhas):
                                # listas de grupos: validar cada produto escolhido
                                for grupo in escolhas:
                                    for escolha_id in grupo:
                                        prod_escolha = Produto.objects.select_for_update().get(id=escolha_id)
                                        if prod_escolha.quantidade < 1:
                                            mensagens_erro.append(
                                                f'Produto "{prod_escolha.nome}" não tem estoque suficiente. '
                                                f'Disponível: {prod_escolha.quantidade}, necessário: 1'
                                            )
                            else:
                                # caso único (ex: escolha por unidade)
                                for escolha_id in escolhas:
                                    prod_escolha = Produto.objects.select_for_update().get(id=escolha_id)
                                    if prod_escolha.quantidade < quantidade:
                                        mensagens_erro.append(
                                            f'Produto "{prod_escolha.nome}" não tem estoque suficiente. '
                                            f'Disponível: {prod_escolha.quantidade}, necessário: {quantidade}'
                                        )

                    else:
                        if produto.quantidade < quantidade:
                            mensagens_erro.append(
                                f'Produto "{produto.nome}" não tem estoque suficiente. '
                                f'Disponível: {produto.quantidade}, necessário: {quantidade}'
                            )

                if mensagens_erro:
                    return JsonResponse({"status": "erro", "mensagens": mensagens_erro})

                # 2. Cria pedido
                pedido = Pedido.objects.create(
                    cliente=cliente,
                    metodo_pagamento=metodo_pagamento,
                    tipo_consumo=tipo_consumo,
                    observacoes=observacoes
                )

                # 3. Cria itens (o save do ItemPedido vai cuidar do estoque)
                for item in itens:
                    produto_id = item.get("produto_id")
                    quantidade = item.get("quantidade", 1)
                    escolhas = item.get("escolhas", [])

                    produto = Produto.objects.get(id=produto_id)

                    if hasattr(produto, "combo") and produto.combo.tipo in ["opcional", "fixo_opcional"]:
                        # trata combos com escolhas (tanto 'opcional' quanto 'fixo_opcional')
                        if isinstance(escolhas, list) and any(isinstance(e, list) for e in escolhas):
                            for grupo in escolhas:
                                ItemPedido.objects.create(
                                    pedido=pedido,
                                    produto=produto,
                                    quantidade=1,
                                    escolhas_combo=grupo
                                )
                        else:
                            # se o usuário escolheu N unidades (quantidade) e forneceu uma lista de escolhas
                            for _ in range(quantidade):
                                ItemPedido.objects.create(
                                    pedido=pedido,
                                    produto=produto,
                                    quantidade=1,
                                    escolhas_combo=escolhas or None
                                )
                    else:
                        ItemPedido.objects.create(
                            pedido=pedido,
                            produto=produto,
                            quantidade=quantidade,
                            escolhas_combo=escolhas or None
                        )



            return JsonResponse({"status": "ok", "pedido_id": pedido.id})

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({"status": "erro", "mensagens": [str(e)]}, status=500)


class EscolherLocalView(View):
    template_name = "pag_escolher_local.html"

    def get(self, request, cliente_id):
        return render(request, self.template_name, {"cliente_id": cliente_id})

    def dispatch(self, request, *args, **kwargs):
        cliente_id = self.kwargs.get('cliente_id')
        cliente = get_object_or_404(Cliente, id=cliente_id)

        pedido_existente = Pedido.objects.filter(cliente=cliente).last()

        if pedido_existente and pedido_existente.liberado_para_caixa:
            print("entrou")
            return redirect('pag_final_cliente', pedido_id=pedido_existente.id)

        return super().dispatch(request, *args, **kwargs)

    def post(self, request, cliente_id):
        tipo_consumo = request.POST.get("tipo_consumo", "").lower()
        observacoes = request.POST.get("observacoes", "")

        if tipo_consumo not in ["local", "viagem"]:
            return JsonResponse({"erro": "Opção inválida"}, status=400)

        cliente = get_object_or_404(Cliente, id=cliente_id)
        pedido = Pedido.objects.filter(cliente=cliente).last()

        if not pedido:
            return JsonResponse({"erro": "Nenhum pedido aberto encontrado"}, status=404)

        pedido.tipo_consumo = tipo_consumo
        pedido.observacoes = observacoes
        pedido.liberado_para_caixa = True
        pedido.save()

        return JsonResponse({"mensagem": "Tipo de consumo salvo com sucesso", "pedido_id": pedido.id})


class PagFinalCliente(TemplateView):
    template_name = "pag_final_cliente.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido_id = self.kwargs.get("pedido_id")
        pedido = get_object_or_404(Pedido, id=pedido_id)
        context["pedido"] = pedido

        itens_processados = []

        for item in pedido.itens.all():
            produto = item.produto
            quantidade = item.quantidade
            nome_base = produto.nome
            combo = getattr(produto, 'combo', None)

            # Combos do tipo 'fixo' (todos os itens fixos listados)
            if combo and combo.tipo == 'fixo':
                nomes_itens_combo = [i.produto.nome for i in combo.itens.all()]
                descricao = f"{nome_base} ({', '.join(nomes_itens_combo)})"
                itens_processados.append({
                    'quantidade': quantidade,
                    'descricao': descricao
                })

            # Combos que possuem opções: tanto 'opcional' quanto 'fixo_opcional'
            elif combo and combo.tipo in ['opcional', 'fixo_opcional']:
                nomes_fixos = [i.produto.nome for i in combo.itens.all()]
                escolhas_ids = item.escolhas_combo or []

                # Normaliza escolhas_ids para lista de ints (suporta tanto ["1","2"] quanto [1,2])
                try:
                    escolhas_ids = [int(e) for e in escolhas_ids] if escolhas_ids else []
                except Exception:
                    escolhas_ids = []

                # pega categorias das escolhas (para checar se há várias categorias)
                categorias_das_escolhas = list(
                    Produto.objects.filter(id__in=escolhas_ids).values_list('categoria', flat=True)
                )
                categorias_unicas = set(categorias_das_escolhas)

                if categorias_unicas and len(categorias_unicas) > 1:
                    # Mais de uma categoria: cada escolha representa 1 unidade do combo (ex: 2 combos)
                    for escolha_id in escolhas_ids:
                        nome_escolha = Produto.objects.filter(id=escolha_id).values_list('nome', flat=True).first()
                        nomes = nomes_fixos.copy()
                        if nome_escolha:
                            nomes.append(nome_escolha)
                        descricao = f"{nome_base} ({', '.join(nomes)})"
                        itens_processados.append({
                            'quantidade': 1,
                            'descricao': descricao
                        })
                else:
                    # Mesma categoria ou nenhuma categoria (fallback)
                    # Se cada escolha corresponde a uma unidade: listar individualmente
                    if escolhas_ids and len(escolhas_ids) == quantidade:
                        for escolha_id in escolhas_ids:
                            nome_escolha = Produto.objects.filter(id=escolha_id).values_list('nome', flat=True).first()
                            nomes = nomes_fixos.copy()
                            if nome_escolha:
                                nomes.append(nome_escolha)
                            descricao = f"{nome_base} ({', '.join(nomes)})"
                            itens_processados.append({
                                'quantidade': 1,
                                'descricao': descricao
                            })
                    else:
                        # fallback agrupado: mostra quantidade X do combo e lista as escolhas concatenadas
                        nomes_escolhas = list(Produto.objects.filter(id__in=escolhas_ids).values_list('nome', flat=True))
                        nomes = nomes_fixos + nomes_escolhas
                        descricao = f"{nome_base} ({', '.join(nomes)})" if nomes else f"{nome_base}"
                        itens_processados.append({
                            'quantidade': quantidade,
                            'descricao': descricao
                        })

            else:
                # Produto comum (não é combo)
                descricao = nome_base
                itens_processados.append({
                    'quantidade': quantidade,
                    'descricao': descricao
                })


        context["itens_processados"] = itens_processados
        return context



@method_decorator(staff_member_required, name='dispatch')
class CozinhaView(View):
    def get(self, request):
        pedidos = Pedido.objects.filter(liberado_para_cozinha=True).order_by( # Só pedidos liberados e ordenados
        # Primeiro por tipo de consumo: viagem antes de local
            Case(
                When(tipo_consumo='viagem', then=Value(0)),
                When(tipo_consumo='local', then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            ),
            'time_liberado_para_cozinha'  # depois por ordem de liberação  
        )
        return render(request, 'pag_cozinha.html', {'pedidos': pedidos})

    def post(self, request):
        data = json.loads(request.body)
        if data.get("acao") == "cancelar":
            try:
                pedido = Pedido.objects.get(pk=data["id"])
                
                # Salvar no modelo PedidoCancelado
                itens_str = " | ".join(f"{item.produto.nome} ({item.quantidade})" for item in pedido.itens.all())
                PedidoCancelado.objects.create(
                    cliente=pedido.cliente,
                    itens=itens_str,
                    metodo_pagamento=pedido.metodo_pagamento
                )

                # Devolver estoque
                for item in pedido.itens.all():
                    # item.produto.quantidade += item.quantidade
                    # item.produto.save()
                    devolver_estoque_do_item(item)

                pedido.delete()  # Excluir pedido original

                return JsonResponse({"status": "success"})
            except Pedido.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Pedido não encontrado"}, status=404)
        
        elif data.get("acao") == "confirmar":
            try:
                pedido = Pedido.objects.get(pk=data["id"])

                # Salvar como entregue
                itens_str = " | ".join(f"{item.produto.nome} ({item.quantidade})" for item in pedido.itens.all())
                PedidoEntregue.objects.create(
                    cliente=pedido.cliente,
                    itens=itens_str,
                    metodo_pagamento=pedido.metodo_pagamento,
                    total=pedido.total  # se aplicável
                )

                pedido.delete()

                return JsonResponse({"status": "success"})
            except Pedido.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Pedido não encontrado"}, status=404)
 
        
        else:
            return JsonResponse({"status": "error", "message": "Ação inválida"}, status=400)       

@method_decorator(staff_member_required, name='dispatch')
class CaixaView(View):
    def get(self, request):
        pedidos = Pedido.objects.filter(liberado_para_cozinha=False, liberado_para_caixa=True)
        return render(request, 'pag_caixa.html', {'pedidos': pedidos})

    def post(self, request):
        data = json.loads(request.body)

        if data.get("acao") == "cancelar":
            try:
                pedido = Pedido.objects.get(pk=data["id"])
                
                # Salvar no modelo PedidoCancelado
                itens_str = " | ".join(f"{item.produto.nome} ({item.quantidade})" for item in pedido.itens.all())
                PedidoCancelado.objects.create(
                    cliente=pedido.cliente,
                    itens=itens_str,
                    metodo_pagamento=pedido.metodo_pagamento
                )

                # Devolver estoque
                for item in pedido.itens.all():
                    # item.produto.quantidade += item.quantidade
                    # item.produto.save()
                    devolver_estoque_do_item(item)

                pedido.delete()

                return JsonResponse({"status": "success"})
            except Pedido.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Pedido não encontrado"}, status=404)
        
           
        else:
            try:
                pedido_id = data.get("pedido_id")

                if not pedido_id:
                    return JsonResponse({'success': False, 'error': 'ID do pedido não encontrado'})

                pedido = get_object_or_404(Pedido, id=pedido_id)

                pedido.liberado_para_cozinha = True
                pedido.time_liberado_para_cozinha = timezone.now()
                pedido.save()

                return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        
def detalhe_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    return render(request, 'pag_detalhe_pedido.html', {'pedido': pedido})

def editar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        observacoes = request.POST.get('observacoes', '')
        pedido.observacoes = observacoes
        pedido.save()

        for item in pedido.itens.all():
            nova_quantidade = int(request.POST.get(f'quantidade_{item.id}', item.quantidade))
            if nova_quantidade != item.quantidade:
                if nova_quantidade == 0:
                    PedidoEntregue.objects.create(
                        cliente=pedido.cliente,
                        itens=f"{item.produto.nome} ({item.quantidade})"
                    )
                    item.delete()
                elif nova_quantidade < item.quantidade:
                    diferenca = item.quantidade - nova_quantidade
                    PedidoEntregue.objects.create(
                        cliente=pedido.cliente,
                        itens=f"{item.produto.nome} ({diferenca})"
                    )
                    item.quantidade = nova_quantidade
                    item.save()
                else:
                    # item.quantidade = nova_quantidade
                    # item.save()
                    diferenca = nova_quantidade - item.quantidade
                    if diferenca > 0:
                        if item.produto.quantidade >= diferenca:
                            item.produto.quantidade -= diferenca
                            item.produto.save()
                            item.quantidade = nova_quantidade
                            item.save()

        return redirect('detalhe_pedido', pedido_id=pedido.id)

    return render(request, 'pag_detalhe_pedido.html', {'pedido': pedido, 'modo_edicao': True})

@staff_member_required
@staff_member_required
def pedidos_json(request):
    pedidos = Pedido.objects.filter(liberado_para_cozinha=True).order_by(
        Case(
            When(tipo_consumo='viagem', then=Value(0)),
            When(tipo_consumo='local', then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        ),
        'time_liberado_para_cozinha'
    )

    lista = []
    for pedido in pedidos:
        produtos_finais = []

        for item in pedido.itens.all():
            produto = item.produto
            quantidade = item.quantidade

            if hasattr(produto, 'combo'):
                combo = produto.combo
                if combo.tipo == 'fixo':
                    for i in combo.itens.all():
                        produtos_finais.extend([i.produto.nome] * quantidade)
                elif combo.tipo == 'opcional':
                    nomes_fixos = [i.produto.nome for i in combo.itens.all()]
                    escolhas_ids = item.escolhas_combo or []
                    escolhas = Produto.objects.filter(id__in=escolhas_ids).values_list('nome', flat=True)

                    produtos_finais.extend(nomes_fixos * quantidade)  # fixos valem 1
                    produtos_finais.extend(escolhas)  # 1 de cada escolha
            else:
                produtos_finais.extend([produto.nome] * quantidade)

        # Agora agrupar os produtos finais
        contador = Counter(produtos_finais)
        itens_formatados = [f"{nome} (x{qtd})" for nome, qtd in contador.items()]


        lista.append({
            'id': pedido.id,
            'cliente': pedido.cliente.nome,
            'itens': itens_formatados,
            'observacoes': pedido.observacoes,
            'tipo_consumo': pedido.tipo_consumo,
        })

    return JsonResponse({'pedidos': lista})


@staff_member_required
@staff_member_required
def pedidos_caixa_json(request):
    pedidos = Pedido.objects.filter(
        liberado_para_cozinha=False,
        liberado_para_caixa=True
    ).order_by('-id').reverse()

    lista = []
    for pedido in pedidos:
        itens_formatados = []

        for item in pedido.itens.all():
            produto = item.produto
            quantidade = item.quantidade

            if hasattr(produto, 'combo'):
                combo = produto.combo
                if combo.tipo == 'fixo':
                    nomes_itens_combo = [i.produto.nome for i in combo.itens.all()]
                    descricao = f"{produto.nome} ({', '.join(nomes_itens_combo)})"
                    itens_formatados.append(f"{descricao} (x{quantidade})")
                elif combo.tipo == 'opcional':
                    nomes_fixos = [i.produto.nome for i in combo.itens.all()]
                    escolhas_ids = item.escolhas_combo or []

                    nomes = nomes_fixos.copy()

                    if escolhas_ids:
                        escolhas_ids = [int(eid) for eid in escolhas_ids]
                        nomes_escolhas = list(
                            Produto.objects.filter(id__in=escolhas_ids).values_list('nome', flat=True)
                        )
                        nomes.extend(nomes_escolhas)

                    descricao = f"{produto.nome} ({', '.join(nomes)})"
                    itens_formatados.append(f"{descricao} (x{quantidade})")

            else:
                descricao = f"{produto.nome}"
                itens_formatados.append(f"{descricao} (x{quantidade})")

        lista.append({
            'id': pedido.id,
            'cliente': pedido.cliente.nome,
            'itens':itens_formatados,
            'observacoes': pedido.observacoes,
            'metodo_pagamento': pedido.metodo_pagamento,
            'tipo_consumo': pedido.tipo_consumo,
            'total': number_format(pedido.total, 2)
        })

    return JsonResponse({'pedidos': lista})


@csrf_exempt
def cancelar_pedido(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pedido_id = data.get('pedido_id')
            pedido = Pedido.objects.get(id=pedido_id)

            if pedido.liberado_para_caixa or pedido.liberado_para_cozinha:
                return JsonResponse({'success': False, 'error': 'Pedido já em andamento, não pode ser cancelado'})

            itens_str = " | ".join([f"{item.produto.nome} ({item.quantidade})" for item in pedido.itens.all()])
            PedidoCancelado.objects.create(
                cliente=pedido.cliente,
                itens=itens_str,
                metodo_pagamento=pedido.metodo_pagamento,
            )

            # Retorna ao estoque
            for item in pedido.itens.all():
                # item.produto.quantidade += item.quantidade
                # item.produto.save()
                devolver_estoque_do_item(item)

            pedido.delete()

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Método não permitido'})

# Garante que apenas admins acessem essa view
@staff_member_required 
def exportar_excel_pedidos(request):
    pedidos = PedidoEntregue.objects.all().order_by('-entregue_em')

    dados = []
    total_dinheiro = 0
    total_pix = 0
    vendas_por_produto = defaultdict(int)

    for pedido in pedidos:
        metodo = pedido.metodo_pagamento.lower()
        total = float(pedido.total)

        if metodo == "dinheiro":
            total_dinheiro += total
        elif metodo == "pix":
            total_pix += total

        itens_texto = str(pedido.itens)
        itens = re.findall(r'([^\(,]+)\s*\(?(\d+)?\)?', itens_texto)

        for nome_item, qtd in itens:
            nome_item = re.sub(r'^[^\w]*|[^\w\s]*$', '', nome_item.strip())
            quantidade = int(qtd) if qtd else 1
            vendas_por_produto[nome_item] += quantidade

        dados.append({
            "Cliente": pedido.cliente.nome,
            "Itens": pedido.itens,
            "Total (R$)": total,
            "Forma de Pagamento": pedido.metodo_pagamento,
        })

    df_pedidos = pd.DataFrame(dados)

    hoje = datetime.now(pytz.timezone('America/Sao_Paulo'))
    data_formatada = hoje.strftime("%d/%m/%Y")
    nome_arquivo = f'Cantina JID - {hoje.strftime("%d/%m")}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        workbook = writer.book
        ws = workbook.create_sheet(title='Relatório')

        ws.sheet_view.showGridLines = False

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        title_font = Font(bold=True, size=14, color="1F4E78")
        destaque_font = Font(bold=True, size=12)
        destaque_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center")
        border = Border(bottom=Side(border_style="thin", color="000000"))

        # Nova borda fina completa
        thin_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        row = 1

        # TÍTULO
        ws.cell(row=row, column=1).value = f"RESUMO DE VENDAS - {data_formatada}"
        ws.cell(row=row, column=1).font = title_font
        ws.cell(row=row, column=1).alignment = left_alignment
        row += 2

        # TOTAL DE VENDAS
        ws.cell(row=row, column=1).value = "TOTAL RECEBIDO EM DINHEIRO:"
        ws.cell(row=row, column=1).font = destaque_font
        ws.cell(row=row, column=1).fill = destaque_fill
        ws.cell(row=row, column=2).value = total_dinheiro
        ws.cell(row=row, column=2).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=2).fill = destaque_fill
        row += 1

        ws.cell(row=row, column=1).value = "TOTAL RECEBIDO EM PIX:"
        ws.cell(row=row, column=1).font = destaque_font
        ws.cell(row=row, column=1).fill = destaque_fill
        ws.cell(row=row, column=2).value = total_pix
        ws.cell(row=row, column=2).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=2).fill = destaque_fill
        row += 1

        ws.cell(row=row, column=1).value = "TOTAL GERAL:"
        ws.cell(row=row, column=1).font = Font(bold=True, size=13, color="006100")
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2).value = total_dinheiro + total_pix
        ws.cell(row=row, column=2).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=2).fill = total_fill
        row += 2

        # PRODUTOS VENDIDOS
        ws.cell(row=row, column=1).value = "Produto"
        ws.cell(row=row, column=2).value = "Quantidade Vendida"
        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = 35
        row += 1

        for produto, qtd in sorted(vendas_por_produto.items()):
            ws.cell(row=row, column=1).value = produto
            ws.cell(row=row, column=2).value = qtd
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.alignment = left_alignment if col == 1 else alignment
                cell.border = thin_border
            row += 1

        row += 2  # espaço antes da tabela de pedidos

        # TABELA DE PEDIDOS
        for col_num, column_title in enumerate(df_pedidos.columns, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_num)].width = 40
        row += 1

        for i, registro in df_pedidos.iterrows():
            for j, value in enumerate(registro, 1):
                cell = ws.cell(row=row + i, column=j)
                cell.value = value
                if j == 3:
                    cell.number_format = 'R$ #,##0.00'
                elif j == 5:
                    cell.number_format = 'DD/MM/YYYY HH:MM'
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border

    return response
